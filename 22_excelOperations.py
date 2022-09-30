import openpyxl
from openpyxl import load_workbook,cell
filename = (r"D:\OneDrive\Desktop\Book1.xlsx")

#locate and open file by file name
workbook = openpyxl.load_workbook(r"D:\OneDrive\Desktop\Book1.xlsx")


##Locate particular sheet
sheet = workbook['S3']

### read data from each cell
# print(sheet.cell(row=1,column=1).value)
# print(sheet.cell(row=1,column=2).value)
# print(sheet.cell(row=2,column=1).value)
# print(sheet.cell(row=2,column=2).value)
#
# ### write in excel
# sheet.cell(row=11,column=1).value="Nihal"
# sheet.cell(row=11,column=2).value="Nihal@256"

workbook.save(filename)

crow = sheet.max_row
print(crow)
ccol = sheet.max_column
print(ccol)


### read all data from file, update multiple records, extract some data to other/new excel

# for m in range ( 1, crow + 1):
#     for n in range ( 1, ccol + 1):
#         print (sheet.cell (row=m, column=n). value)

###################################################################################################

for i in range(1,crow+1):
    for j in range(1,ccol+1):
        print(sheet.cell(i,j).value,end="    ")
    print()


###################################################################################################

# for m in range ( 1, crow + 1):
# # to get all the cell values from column 1
#     if sheet.cell (row=m, column=1).value == "abhinav" :
# # iterate till the count of occupied columns
#         for n in range ( 1, ccol + 1):
# # to get all the cell data from a row and print
#             print (sheet.cell (row=m, column=n). value)
#
#
#
#









